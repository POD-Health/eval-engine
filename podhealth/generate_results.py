import json
import os
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from provider import _get_token, _parse_sse

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

excel_path = "src/docs/testfile.xlsx"
result_path = "podhealth/result.json"
model_version = "Piper AI Agent"

wb = load_workbook(excel_path)
ws = wb.active

results = []


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def calculate_score(expected, actual_response):
    expected_lines = [
        line.strip("-• \n").strip()
        for line in str(expected).splitlines()
        if line.strip()
    ]

    lowered_response = clean_text(actual_response).lower()
    matched = 0

    for point in expected_lines:
        point_lower = point.lower()
        key_words = [w for w in point_lower.split() if len(w) > 3]

        if point_lower in lowered_response or any(word in lowered_response for word in key_words):
            matched += 1

    total = max(len(expected_lines), 1)
    score = round((matched / total) * 10)

    if matched == total:
        reason = "Response covered all expected points."
    elif matched == 0:
        reason = "Response did not cover the expected points."
    else:
        reason = f"Response covered {matched} out of {total} expected points."

    return score, reason


def get_actual_response(question):
    debug_info = {
        "url": "",
        "question": question,
        "payload": {},
        "status_code": None,
        "retry_status_code": None,
        "raw_response_preview": "",
        "error_stage": "",
    }

    try:
        token = _get_token()

        base_url = os.environ["DATA_AGENT_BASE_URL"].rstrip("/")
        diagnostic_id = os.environ["PARENT_DIAGNOSTIC_ID"]
        child_id = os.environ.get("CHILD1_ID", "")

        url = f"{base_url}/agno-query-sql-agent"
        debug_info["url"] = url

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }

        payload = {
            "stream": True,
            "diagnostic_ids": [diagnostic_id],
            "stakeholder": "parent",
            "question": f"For patient {child_id}: {question}",
            "conversation_id": f"eval_{int(datetime.now().timestamp())}",
            "current_date": datetime.now().strftime("%Y-%m-%d"),
            "timezone": "UTC"
        }
        debug_info["payload"] = payload

        response = requests.post(
            url,
            json=payload,
            headers=headers,
            stream=True,
            timeout=120
        )

        debug_info["status_code"] = response.status_code

        if response.status_code == 401:
            debug_info["error_stage"] = "initial_request_401"

            token = _get_token()
            headers["Authorization"] = f"Bearer {token}"

            response = requests.post(
                url,
                json=payload,
                headers=headers,
                stream=True,
                timeout=120
            )

            debug_info["retry_status_code"] = response.status_code

        if response.status_code != 200:
            try:
                raw_text = response.text
            except Exception:
                raw_text = ""

            debug_info["raw_response_preview"] = raw_text[:1000]
            return f"API Error {response.status_code}: {raw_text}", debug_info

        text = _parse_sse(response)

        if not text:
            debug_info["error_stage"] = "empty_sse_response"
            return "Piper returned an empty response", debug_info

        debug_info["raw_response_preview"] = text[:1000]
        return text, debug_info

    except Exception as e:
        debug_info["error_stage"] = "exception"
        debug_info["raw_response_preview"] = f"{type(e).__name__}: {e}"
        return f"Request failed: {type(e).__name__}: {e}", debug_info


for row in range(5, ws.max_row + 1):
    question_id = clean_text(ws.cell(row=row, column=2).value)   
    question = clean_text(ws.cell(row=row, column=3).value)      
    expected = clean_text(ws.cell(row=row, column=4).value)      

    if not question_id or not question or not expected:
        continue

    actual_response, debug_info = get_actual_response(question)
    score, reason = calculate_score(expected, actual_response)

    results.append({
        "question_id": question_id,
        "question": question,
        "expected_answer": expected,
        "response": actual_response,
        "score": score,
        "reason": reason,
        "response_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "model_version": model_version,
        "debug": debug_info
    })

with open(result_path, "w", encoding="utf-8") as f:
    json.dump(results, f, indent=2, ensure_ascii=False)

print(f"Generated {len(results)} results in {result_path}")