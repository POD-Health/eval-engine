from openpyxl import load_workbook
import json
from datetime import datetime

file_path = "src/docs/testfile.xlsx"
promptfoo_result_path = "podhealth/promptfoo-results.json"

wb = load_workbook(file_path)

with open(promptfoo_result_path, "r", encoding="utf-8") as f:
    data = json.load(f)

results = data.get("results", {}).get("results", [])

result_map = {}

for item in results:
    if not isinstance(item, dict):
        continue

    test_case = item.get("testCase", {})
    if not isinstance(test_case, dict):
        test_case = {}

    description = str(test_case.get("description", "")).strip()
    qid = ""

    # Example:
    # Q-001 [General health overview] Can you tell me...
    if description.startswith("Q-"):
        qid = description.split()[0].strip()

    vars_data = item.get("vars", {})
    if not isinstance(vars_data, dict):
        vars_data = {}

    question = str(vars_data.get("question", "")).strip()

    response_data = item.get("response", {})
    if isinstance(response_data, dict):
        actual_response = str(response_data.get("output", "")).strip()
    else:
        actual_response = str(response_data).strip()

    grading_result = item.get("gradingResult", {})
    if not isinstance(grading_result, dict):
        grading_result = {}

    raw_score = grading_result.get("score", 0)
    try:
        raw_score = float(raw_score)
    except Exception:
        raw_score = 0.0

    score = round(raw_score * 10)
    passed = grading_result.get("pass", False)
    reason = str(grading_result.get("reason", "")).strip()

    pass_fail = "Pass" if passed else "Fail"

    if qid:
        result_map[qid] = {
            "question": question,
            "response": actual_response,
            "score": score,
            "pass_fail": pass_fail,
            "reason": reason,
        }

excluded_sheets = ["Score Summary", "Instructions", "Version Log"]
target_sheets = [sheet for sheet in wb.sheetnames if sheet not in excluded_sheets]

updated_count = 0

for sheet_name in target_sheets:
    ws = wb[sheet_name]

    header_row_idx = None
    qid_col = None

    # Find header row dynamically
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(str(cell_val).strip() if cell_val is not None else "")

        if "Question ID" in row_values:
            header_row_idx = row_idx
            qid_col = row_values.index("Question ID") + 1
            break

    if header_row_idx is None or qid_col is None:
        print(f"Skipping sheet without Question ID header: {sheet_name}")
        continue

    for row in range(header_row_idx + 1, ws.max_row + 1):
        qid = str(ws.cell(row=row, column=qid_col).value or "").strip()

        if not qid or not qid.startswith("Q-"):
            continue

        matched_item = result_map.get(qid)
        if not matched_item:
            continue

        actual_response = matched_item["response"]
        score = matched_item["score"]
        pass_fail = matched_item["pass_fail"]
        reason = matched_item["reason"]

        accuracy = f"{score * 10}%"

        if pass_fail == "Pass":
            retest = "No"
            priority = "Low"
            tester_notes = "Promptfoo marked this as passing"
        else:
            retest = "Yes"
            priority = "High"
            tester_notes = "Promptfoo marked this as failing"

        # F to O columns
        ws.cell(row=row, column=6).value = actual_response
        ws.cell(row=row, column=7).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row, column=8).value = "Piper AI Agent"
        ws.cell(row=row, column=9).value = score
        ws.cell(row=row, column=10).value = accuracy
        ws.cell(row=row, column=11).value = pass_fail
        ws.cell(row=row, column=12).value = reason if reason else "Taken from Promptfoo output"
        ws.cell(row=row, column=13).value = tester_notes
        ws.cell(row=row, column=14).value = retest
        ws.cell(row=row, column=15).value = priority

        updated_count += 1

wb.save(file_path)
print(f"Excel updated successfully from Promptfoo results. Total rows updated: {updated_count}")