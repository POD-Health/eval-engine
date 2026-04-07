"""
export_results.py - Convert promptfoo eval JSON output to a formatted Excel report.

Usage:
    1. Run your eval with JSON output:
       npm run local -- eval -c podhealth/promptfooconfig.yaml --env-file podhealth/.env --no-cache -o podhealth/results.json

    2. Then run this script:
       python podhealth/export_results.py

Output: podhealth/eval_results.xlsx
"""

import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "podhealth/results.json"
OUTPUT_FILE = "podhealth/eval_results.xlsx"

# Colors
COLOR_HEADER_BG = "1F4E79"   # Dark blue header
COLOR_HEADER_FG = "FFFFFF"   # White text
COLOR_PASS_BG   = "C6EFCE"   # Green pass
COLOR_FAIL_BG   = "FFCCCC"   # Red fail
COLOR_ERROR_BG  = "FFE699"   # Yellow error
COLOR_ALT_ROW   = "F2F7FF"   # Light blue alternating row

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def load_results(path):
    with open(path) as f:
        data = json.load(f)
    return data

def parse_rows(data):
    rows = []
    results = data.get("results", {}).get("results", data.get("results", []))
    if isinstance(results, dict):
        results = results.get("results", [])

    for r in results:
        question = ""
        rubric = ""
        response = ""
        status = "error"
        score = 0.0

        # Extract vars
        vars_ = r.get("vars", {})
        question = vars_.get("question", "")
        rubric = vars_.get("rubric", "")

        # Extract response
        response_obj = r.get("response", {})
        if response_obj:
            output = response_obj.get("output", "")
            if isinstance(output, str):
                response = output
            elif isinstance(output, dict):
                response = output.get("text", str(output))

        # Extract pass/fail/score
        success = r.get("success", False)
        score = r.get("score", 0.0)
        error = r.get("error", "")

        if error:
            status = "error"
        elif success:
            status = "pass"
        else:
            status = "fail"

        rows.append({
            "question": question,
            "rubric": rubric,
            "response": response,
            "status": status,
            "score": score,
            "error": error,
        })

    return rows

def write_excel(rows, output_path):
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────
    summary = wb.active
    summary.title = "Summary"

    total  = len(rows)
    passed = sum(1 for r in rows if r["status"] == "pass")
    failed = sum(1 for r in rows if r["status"] == "fail")
    errors = sum(1 for r in rows if r["status"] == "error")
    avg_score = (sum(r["score"] for r in rows) / total) if total else 0

    summary_data = [
        ["PodHealth Piper AI — Eval Results"],
        [],
        ["Metric",          "Value"],
        ["Total Questions", total],
        ["Passed",          passed],
        ["Failed",          failed],
        ["Errors",          errors],
        ["Pass Rate",       f"=D5/D4" if total else "0%"],
        ["Avg Score",       round(avg_score, 3)],
    ]

    for row in summary_data:
        summary.append(row)

    # Title style
    summary["A1"].font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    summary["A1"].fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    summary.merge_cells("A1:B1")
    summary["A1"].alignment = Alignment(horizontal="center")

    # Header row
    for cell in summary[3]:
        cell.font = Font(name="Arial", bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)

    # Color the pass/fail/error counts
    color_map = {"D5": COLOR_PASS_BG, "D6": COLOR_FAIL_BG, "D7": COLOR_ERROR_BG}
    for addr, color in color_map.items():
        summary[addr].fill = PatternFill("solid", start_color=color)

    # Format pass rate as percentage
    summary["D8"].number_format = "0.0%"

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14

    # ── Results sheet ──────────────────────────────────────────────
    ws = wb.create_sheet("Results")

    headers = ["#", "Status", "Score", "Question", "Expected (Rubric)", "Piper Response", "Error"]
    ws.append(headers)

    # Header styling
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name="Arial", bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()

    # Data rows
    for i, row in enumerate(rows, 1):
        excel_row = i + 1
        status = row["status"]

        bg = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        if status == "pass":
            bg = COLOR_PASS_BG
        elif status == "fail":
            bg = COLOR_FAIL_BG
        elif status == "error":
            bg = COLOR_ERROR_BG

        values = [
            i,
            status.upper(),
            row["score"],
            row["question"],
            row["rubric"],
            row["response"],
            row["error"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border()

        # Score formatting
        ws.cell(row=excel_row, column=3).number_format = "0.00"

    # Column widths
    ws.column_dimensions["A"].width = 5   # #
    ws.column_dimensions["B"].width = 10  # Status
    ws.column_dimensions["C"].width = 8   # Score
    ws.column_dimensions["D"].width = 40  # Question
    ws.column_dimensions["E"].width = 45  # Rubric
    ws.column_dimensions["F"].width = 55  # Response
    ws.column_dimensions["G"].width = 30  # Error

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:G{len(rows)+1}"

    wb.save(output_path)
    print(f"✅ Excel report saved to: {output_path}")
    print(f"   {passed}/{total} passed ({100*passed//total if total else 0}%)")

def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    output_file = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_FILE

    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        print(f"   Run your eval with: npm run local -- eval -c podhealth/promptfooconfig.yaml --env-file podhealth/.env --no-cache -o {input_file}")
        sys.exit(1)

    print(f"📂 Loading results from: {input_file}")
    data = load_results(input_file)
    rows = parse_rows(data)
    print(f"   Found {len(rows)} test results")
    write_excel(rows, output_file)

if __name__ == "__main__":
    main()